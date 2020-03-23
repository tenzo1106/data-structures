import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
import xlrd
import matplotlib.pyplot as plt
def View():
     print("VIEWING DATA : ")
     print("Enter (1)Year  (2)Crop")
     d=int(input())
     if d==1 :
        year = int(input("Enter year: "))
        for sh in xlrd.open_workbook(r'C:\Users\Lenovo\Desktop\dwd\crops.xlsx').sheets():  
            for row in range(sh.nrows):
                for col in range(sh.ncols):
                    myCell = sh.cell(row, col)
                    if myCell.value == year:
                        print('-----------')
                                    #print('Found!')
                        l = xl_rowcol_to_cell(row,col)
                        x = list(l)
                                    #print(x)
                        al = x[1:]
                                    #print(al)
                                    #xl = int(str(al))
                                    #print(al)
                        t=''
                        for j in al:
                             t=t+j
                             #print(t)
                             t2 = int(t)
                                    #print(xl_rowcol_to_cell(row,col))
                             print(sh.row_values(t2-1))
                                    #for i in range(sh.ncols):
                                    #   print(sh.cell_value(0, i))               

     elif d==2:
        cropname = input("Enter cropname: ")                
        Alist=[]
        Blist=[]
        for sh in xlrd.open_workbook(r'C:\Users\Lenovo\Desktop\dwd\crops.xlsx').sheets():  
            for row in range(sh.nrows):
                for col in range(sh.ncols):
                    myCell = sh.cell(row, col)
                    if myCell.value == cropname.title():
                        print('-----------')
                        l = xl_rowcol_to_cell(row,col)
                        x = list(l)                                
                        al = x[1:]                               
                        t=''
                        for j in al:
                            t=t+j                                
                        t2 = int(t)                                
                        print(sh.row_values(t2-1))
                        Alist.append(sh.row_values(t2-1)[2])
                        Blist.append(sh.row_values(t2-1)[6])
        import matplotlib.pyplot as plt; plt.rcdefaults()
        import numpy as np
        import matplotlib.pyplot as plt

        objects =Alist
        performance =Blist

        plt.bar(objects, performance, align='center')
        
        plt.ylabel('production')
        plt.title('years')

        plt.show()


while True :
    f1=0
    print("Enter : (1)Admin (2)User (3)Exit")
    a=int(input())

    if a==1 :
        print("Welcome Admin")
        print("Enter password : ",end="")
        pw=str(input())
        if pw=="1234" :
            print("Loging in...")
        else :
            print("Wrong password")
            continue
        while True :
            print("Enter : (1)add (2)View (3)Back to Main menu (4)Exit ")
            b=int(input())
            if b==1 :
                print("ADDING DATA : ")
                print("Enter State : ",end="")
                state=input()
                print("Enter District : ",end="")
                district=input()
                print("Enter Crop year : ",end="")
                year=int(input())
                print("Enter Season : ",end="")
                season=input()
                print("Enter Crop : ",end="")
                crop=input()
                print("Enter Area : ",end="")
                area=float(input())
                print("Enter Production : ",end="")
                prod=float(input())
                df = pd.DataFrame({'State_Name': [state.title()],
                                    'District_Name': [district.upper()],
                                    'Crop_Year':[year],
                                    'Season':[season.title()],
                                    'Crop':[crop.title()],
                                    'Area':[area],
                                    'Production':[prod]})
                writer = pd.ExcelWriter(r'C:\Users\Lenovo\Desktop\dwd\crops.xlsx', engine='openpyxl')
                writer.book = load_workbook(r'C:\Users\Lenovo\Desktop\dwd\crops.xlsx')
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                reader = pd.read_excel(r'C:\Users\Lenovo\Desktop\dwd\crops.xlsx')
                df.to_excel(writer,index=False,header=False,startrow=len(reader)+1)
                writer.close()
                print("Data added successfully.")
                print("Enter : (1)Going back (2)Going back to main menu (3)Exit")
                c=int(input())
                if c==1 :
                    print("Going back...")
                    continue
                elif c==2 :
                    print("Going back to main menu...")
                    break
                elif c==3 :
                    print("Thank you")
                    print("Visit agin.")
                    f1=1
                    break
            elif b==2 :
                View()
                print("Enter : (1)Going back (2)Going back to main menu (3)Exit")
                c=int(input())
                if c==1 :
                    print("Going back...")
                    continue
                elif c==2 :
                    print("Going back to main menu...")
                    break
                elif c==3 :
                    print("Thank you")
                    print("Visit agin.")
                    f1=1
                    break
            elif b==3 :
                print("Going back to main menu...")
                break

            elif b==4 :
                print("Thank you")
                print("Visit agin.")
                f1=1
                break


    elif a==2 :
        print("Hello User")
        View()
        print("Enter : (1)Going back to main menu (2)Exit")
        c=int(input())
        if c==1 :
            print("Going back to main menu...")
            continue
        elif c==2 :
            print("Thank you")
            print("Visit agin.")
            break
        

    elif a==3 :
        print("Thank you")
        print("Visit agin.")
        break

    if f1==1 :
        break